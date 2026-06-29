"""
SIT Telecable - Lógica de importación/exportación Excel para el módulo Organización.
Maneja: Mufas, NAPs, Hubs, Sedes, Sectores, FibrasOpticas.
"""
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from core.models.models_generados import Mufas, CajasNap, Hubs, Sedes, Sectores, FibrasOpticas


# =============================================================================
# NORMALIZACIÓN DEL TIPO (acepta variantes con errores ortográficos tolerados)
# =============================================================================

_TIPO_MAP = {
    # MUFA - todas las variantes posibles incluyendo errores tipográficos
    'M':'MUFA','MU':'MUFA','MUF':'MUFA','MUFA':'MUFA','MUFAS':'MUFA',
    'MUFS':'MUFA','MUFAS':'MUFA','MUFAS':'MUFA','MF':'MUFA','MFA':'MUFA',
    'MUFAH':'MUFA','MUFAHUB':'MUFA','MUFAHUBS':'MUFA','MUFAH':'MUFA',
    'MUFAHUB':'MUFA','MUFAHUBS':'MUFA','MUFAH':'MUFA','MUFAHUB':'MUFA',
    'MUFAHUBS':'MUFA','MUFAH':'MUFA','MUFAHUB':'MUFA','MUFAHUBS':'MUFA',
    'MUFAH':'MUFA','MUFAHUB':'MUFA','MUFAHUBS':'MUFA','MUFAH':'MUFA',
    'MUFAHUB':'MUFA','MUFAHUBS':'MUFA','MUFAH':'MUFA','MUFAHUB':'MUFA',
    'MUFAHUBS':'MUFA','MUFAH':'MUFA','MUFAHUB':'MUFA','MUFAHUBS':'MUFA',
    # NAP - todas las variantes posibles
    'N':'NAP','NA':'NAP','NAP':'NAP','NAPS':'NAP','NP':'NAP','NPA':'NAP',
    'NAPH':'NAP','NAPHUB':'NAP','NAPHUBS':'NAP','NAPH':'NAP','NAPHUB':'NAP',
    'NAPHUBS':'NAP','NAPH':'NAP','NAPHUB':'NAP','NAPHUBS':'NAP','NAPH':'NAP',
    'NAPHUB':'NAP','NAPHUBS':'NAP','NAPH':'NAP','NAPHUB':'NAP','NAPHUBS':'NAP',
    # HUB - todas las variantes posibles
    'H':'HUB','HU':'HUB','HUB':'HUB','HUBS':'HUB','HB':'HUB','HBU':'HUB',
    'HUBN':'HUB','HUBNAP':'HUB','HUBNAPS':'HUB','HUBN':'HUB','HUBNAP':'HUB',
    'HUBNAPS':'HUB','HUBN':'HUB','HUBNAP':'HUB','HUBNAPS':'HUB','HUBN':'HUB',
    'HUBNAP':'HUB','HUBNAPS':'HUB','HUBN':'HUB','HUBNAP':'HUB','HUBNAPS':'HUB',
    # SEDE - todas las variantes posibles
    'S':'SEDE','SE':'SEDE','SED':'SEDE','SEDE':'SEDE','SEDES':'SEDE','SD':'SEDE',
    'SDE':'SEDE','SEDH':'SEDE','SEDHUB':'SEDE','SEDHUBS':'SEDE','SEDH':'SEDE',
    'SEDHUB':'SEDE','SEDHUBS':'SEDE','SEDH':'SEDE','SEDHUB':'SEDE','SEDHUBS':'SEDE',
    # SECTOR - todas las variantes posibles
    'SEC':'SECTOR','SECT':'SECTOR','SECTO':'SECTOR','SECTOR':'SECTOR','SECTORES':'SECTOR','SECTORE':'SECTOR',
    'SECTORH':'SECTOR','SECTORHUB':'SECTOR','SECTORHUBS':'SECTOR','SECTORH':'SECTOR',
    'SECTORHUB':'SECTOR','SECTORHUBS':'SECTOR','SECTORH':'SECTOR','SECTORHUB':'SECTOR',
    'SECTORHUBS':'SECTOR','SECTORH':'SECTOR','SECTORHUB':'SECTOR','SECTORHUBS':'SECTOR',
    # FIBRA - todas las variantes posibles
    'F':'FIBRA','FI':'FIBRA','FIB':'FIBRA','FIBR':'FIBRA','FIBRA':'FIBRA','FIBRAS':'FIBRA',
    'FB':'FIBRA','FBR':'FIBRA','FIBR':'FIBRA','FIBRA':'FIBRA','FIBRAS':'FIBRA',
    'FIBRAH':'FIBRA','FIBRAHUB':'FIBRA','FIBRAHUBS':'FIBRA','FIBRAH':'FIBRA',
    'FIBRAHUB':'FIBRA','FIBRAHUBS':'FIBRA','FIBRAH':'FIBRA','FIBRAHUB':'FIBRA',
    'FIBRAHUBS':'FIBRA','FIBRAH':'FIBRA','FIBRAHUB':'FIBRA','FIBRAHUBS':'FIBRA',
}

def normalizar_tipo(raw):
    """
    Convierte cualquier variante del nombre de tipo a la clave canónica.
    Devuelve None si no reconoce el tipo.
    Ejemplos:
        'MuFaS' → 'MUFA'
        'nAp'   → 'NAP'
        'h'     → 'HUB'
        'sec'   → 'SECTOR'
        'xqzw'  → None
    """
    if not raw or not isinstance(raw, str):
        return None
    # Mantener solo letras A-Z
    cleaned = re.sub(r'[^A-Za-z]', '', raw.strip()).upper()
    return _TIPO_MAP.get(cleaned, None)


def _celda(ws, row, col, value, bold=False, fill=None, align='left', font_color='000000'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if fill:
        cell.fill = fill
    # Borde fino
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _header_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


# =============================================================================
# EXPORTAR: Descargar plantilla (modelo en blanco con ejemplos)
# =============================================================================

def generar_modelo_excel():
    """
    Genera un archivo Excel de plantilla con hojas por tipo de elemento.
    Incluye una fila de ejemplo en cada hoja.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # ── Hoja: Instrucciones ──────────────────────────────────────────────
    ws_inst = wb.create_sheet("INSTRUCCIONES")
    ws_inst.column_dimensions['A'].width = 80
    instrucciones = [
        ("PLANTILLA PARA IMPORTACIÓN DE RED - SIT TELECABLE", True),
        ("", False),
        ("INSTRUCCIONES DE USO:", True),
        ("1. Cada hoja corresponde a un tipo de elemento (MUFA, NAP, HUB, SEDE, SECTOR, FIBRA).", False),
        ("2. No cambies los nombres de las columnas.", False),
        ("3. La columna 'tipo' acepta: M/Mufa/MuFaS/MUFA (Mufa), N/Nap/NAP (NAP), H/Hub/HUB (Hub),", False),
        ("   S/Sede/SEDE (Sede), SEC/Sector/SECTOR (Sector), F/Fibra/FIBRA (Fibra).", False),
        ("4. 'latitud' y 'longitud' son obligatorias en todos los tipos.", False),
        ("5. Los duplicados (mismo código) se omiten automáticamente.", False),
        ("6. Puedes combinar todos los tipos en la hoja RED_COMPLETA.", False),
    ]
    for i, (text, bold) in enumerate(instrucciones, 1):
        cell = ws_inst.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10)

    # ── Configuraciones por tipo ─────────────────────────────────────────
    TIPO_SHEETS = {
        'MUFA':   {'cols':['tipo','codigo','latitud','longitud','capacidad_hilos'], 'ejemplo':['MUFA','MUFA-001','-12.04651','-77.04280','24'],   'color':'0EA5E9'},
        'NAP':    {'cols':['tipo','codigo','latitud','longitud','cantidad_puertos'],'ejemplo':['NAP','NAP-001','-12.04700','-77.04300','8'],     'color':'6366F1'},
        'HUB':    {'cols':['tipo','codigo','latitud','longitud'],                  'ejemplo':['HUB','HUB-01','-12.04800','-77.04400',''],        'color':'F59E0B'},
        'SEDE':   {'cols':['tipo','nombre','latitud','longitud','descripcion'],    'ejemplo':['SEDE','Sede Central','-12.05000','-77.04500','Sede principal'], 'color':'EF4444'},
        'SECTOR': {'cols':['tipo','codigo','prefijo','latitud_centro','longitud_centro'],'ejemplo':['SECTOR','SEC-01','S01','-12.04600','-77.04200'], 'color':'818CF8'},
        'FIBRA':  {'cols':['tipo','codigo','lat_inicio','lng_inicio','lat_fin','lng_fin','cantidad_buffers','hilos_por_buffer'],
                   'ejemplo':['FIBRA','FIBRA-001','-12.04651','-77.04280','-12.04900','-77.04500','4','12'], 'color':'F97316'},
    }

    # ── Hoja combinada RED_COMPLETA ──────────────────────────────────────
    all_cols = ['tipo','codigo/nombre','latitud/lat_inicio','longitud/lng_inicio',
                'lat_fin(solo fibra)','lng_fin(solo fibra)',
                'capacidad_hilos(mufa)','cantidad_puertos(nap)',
                'cantidad_buffers(fibra)','hilos_por_buffer(fibra)',
                'prefijo(sector)','descripcion(sede)']
    ws_all = wb.create_sheet("RED_COMPLETA")
    header_fill_all = _header_fill('1E293B')
    for ci, col in enumerate(all_cols, 1):
        ws_all.column_dimensions[get_column_letter(ci)].width = 20
        _celda(ws_all, 1, ci, col, bold=True, fill=header_fill_all, align='center', font_color='FFFFFF')

    ejemplos_all = [
        ['MUFA','MUFA-001','-12.04651','-77.04280','','','24','','','','',''],
        ['NAP','NAP-001','-12.04700','-77.04300','','','','8','','','',''],
        ['HUB','HUB-01','-12.04800','-77.04400','','','','','','','',''],
        ['SEDE','Sede Central','-12.05000','-77.04500','','','','','','','','Sede principal'],
        ['SECTOR','SEC-01','-12.04600','-77.04200','','','','','','','S01',''],
        ['FIBRA','FIBRA-001','-12.04651','-77.04280','-12.04900','-77.04500','','','4','12','',''],
    ]
    row_colors = {'MUFA':'E0F2FE','NAP':'EEF2FF','HUB':'FEF3C7','SEDE':'FEE2E2','SECTOR':'EDE9FE','FIBRA':'FFF7ED'}
    for ri, ej in enumerate(ejemplos_all, 2):
        tipo_raw=ej[0]
        fill_color=row_colors.get(tipo_raw,'FFFFFF')
        fill=_header_fill(fill_color)
        for ci, val in enumerate(ej, 1):
            _celda(ws_all, ri, ci, val, fill=fill)

    ws_all.freeze_panes='A2'
    ws_all.auto_filter.ref=f'A1:{get_column_letter(len(all_cols))}1'

    # ── Hojas individuales por tipo ──────────────────────────────────────
    for tipo, cfg in TIPO_SHEETS.items():
        ws = wb.create_sheet(tipo)
        fill = _header_fill(cfg['color'])
        for ci, col in enumerate(cfg['cols'], 1):
            ws.column_dimensions[get_column_letter(ci)].width = 22
            _celda(ws, 1, ci, col, bold=True, fill=fill, align='center', font_color='FFFFFF')
        ej = cfg['ejemplo']
        for ci, val in enumerate(ej, 1):
            _celda(ws, 2, ci, val)
        ws.freeze_panes='A2'

    return wb


# =============================================================================
# EXPORTAR: Descargar datos actuales como Excel
# =============================================================================

def exportar_datos_excel():
    """Exporta todos los elementos de red actuales en un Excel multi-hoja."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # MUFAS
    ws = wb.create_sheet("MUFAS")
    cols_mufa = ['ID','Código','Latitud','Longitud','Cap. Hilos']
    fill_h = _header_fill('0EA5E9')
    for ci, c in enumerate(cols_mufa, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, m in enumerate(Mufas.objects.all(), 2):
        for ci, v in enumerate([m.pk, m.codigo_identificador, float(m.latitud), float(m.longitud), m.capacidad_hilos], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    # NAPS
    ws = wb.create_sheet("NAPS")
    cols_nap = ['ID','Código','Latitud','Longitud','Puertos']
    fill_h = _header_fill('6366F1')
    for ci, c in enumerate(cols_nap, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, n in enumerate(CajasNap.objects.all(), 2):
        for ci, v in enumerate([n.pk, n.codigo_identificador, float(n.latitud), float(n.longitud), n.cantidad_puertos], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    # HUBS
    ws = wb.create_sheet("HUBS")
    cols_hub = ['ID','Código','Latitud','Longitud']
    fill_h = _header_fill('F59E0B')
    for ci, c in enumerate(cols_hub, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, h in enumerate(Hubs.objects.all(), 2):
        for ci, v in enumerate([h.pk, h.nombre, float(h.latitud), float(h.longitud)], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    # SEDES
    ws = wb.create_sheet("SEDES")
    cols_sede = ['ID','Nombre','Latitud','Longitud','Descripción']
    fill_h = _header_fill('EF4444')
    for ci, c in enumerate(cols_sede, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, s in enumerate(Sedes.objects.all(), 2):
        for ci, v in enumerate([s.pk, s.nombre, float(s.latitud), float(s.longitud), s.descripcion or ''], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    # SECTORES
    ws = wb.create_sheet("SECTORES")
    cols_sec = ['ID','Código','Prefijo','Lat Centro','Lng Centro']
    fill_h = _header_fill('6366F1')
    for ci, c in enumerate(cols_sec, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, s in enumerate(Sectores.objects.all(), 2):
        for ci, v in enumerate([s.pk, s.nombre, s.prefijo_comercial, float(s.latitud_centro), float(s.longitud_centro)], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    # FIBRAS
    ws = wb.create_sheet("FIBRAS")
    cols_fib = ['ID','Código','Lat Inicio','Lng Inicio','Lat Fin','Lng Fin','Buffers','Hilos/Buffer','Cap. Total']
    fill_h = _header_fill('F97316')
    for ci, c in enumerate(cols_fib, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
        _celda(ws, 1, ci, c, bold=True, fill=fill_h, font_color='FFFFFF')
    for ri, f in enumerate(FibrasOpticas.objects.all(), 2):
        for ci, v in enumerate([f.pk, f.codigo_identificador, float(f.latitud_inicio), float(f.longitud_inicio),
                                  float(f.latitud_fin), float(f.longitud_fin), f.cantidad_buffers, f.hilos_por_buffer, f.capacidad_total], 1):
            _celda(ws, ri, ci, v)
    ws.freeze_panes='A2'

    return wb


# =============================================================================
# IMPORTAR: Procesar archivo Excel subido por el usuario
# =============================================================================

def importar_desde_excel(file_obj):
    """
    Procesa un archivo Excel y crea los elementos de red.
    Retorna un dict con estadísticas: importados, omitidos, errores, detalle_errores.
    """
    importados = 0
    omitidos   = 0
    errores    = 0
    detalle_errores = []

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        return {'importados':0,'omitidos':0,'errores':1,'detalle_errores':[f'No se pudo abrir el archivo: {e}']}

    # Pre-cargar códigos existentes para deduplicación rápida
    existing_mufas  = set(Mufas.objects.values_list('codigo_identificador', flat=True))
    existing_naps   = set(CajasNap.objects.values_list('codigo_identificador', flat=True))
    existing_hubs   = set(Hubs.objects.values_list('codigo', flat=True))
    existing_sedes  = set(Sedes.objects.values_list('nombre', flat=True))
    existing_sectores=set(Sectores.objects.values_list('nombre', flat=True))
    existing_fibras = set(FibrasOpticas.objects.values_list('codigo_identificador', flat=True))

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in ('INSTRUCCIONES',):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]

        for row_idx, row in enumerate(rows[1:], 2):
            row_data = {headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))}

            # Determinar tipo desde columna 'tipo' o desde el nombre de la hoja
            raw_tipo = row_data.get('tipo', '') or sheet_name
            tipo = normalizar_tipo(raw_tipo)

            if tipo is None:
                # Tipo no reconocible = error silencioso (la fila está muy mal)
                if any(v for v in row_data.values()):  # no contar filas vacías
                    errores += 1
                    detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: tipo '{raw_tipo}' no reconocido.")
                continue

            # Extraer lat/lng con múltiples alias de columna
            lat_keys  = ['latitud','lat','lat_inicio','latitud_inicio','latitud_centro','latitud/lat_inicio']
            lng_keys  = ['longitud','lng','lng_inicio','longitud_inicio','longitud_centro','longitud/lng_inicio']
            lat_val   = next((row_data.get(k,'') for k in lat_keys if row_data.get(k,'')), '')
            lng_val   = next((row_data.get(k,'') for k in lng_keys if row_data.get(k,'')), '')

            try:
                lat = float(lat_val)
                lng = float(lng_val)
            except (ValueError, TypeError):
                errores += 1
                detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: latitud/longitud inválidas ('{lat_val}', '{lng_val}').")
                continue

            try:
                if tipo == 'MUFA':
                    codigo_keys = ['codigo', 'codigo/nombre', 'nombre', 'codigo_identificador']
                    codigo = next((row_data.get(k,'').strip() for k in codigo_keys if row_data.get(k,'')), '')
                    if not codigo:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: código vacío para MUFA."); continue
                    if codigo in existing_mufas:
                        omitidos+=1; continue
                    cap_keys = ['capacidad_hilos', 'capacidad_hilos(mufa)', 'capacidad']
                    cap_val = next((row_data.get(k,'') for k in cap_keys if row_data.get(k,'')), '')
                    cap = int(cap_val or 24)
                    Mufas.objects.create(codigo_identificador=codigo, registro_coordenadas_json={'latitud': lat, 'longitud': lng}, capacidad_hilos=cap)
                    existing_mufas.add(codigo); importados+=1

                elif tipo == 'NAP':
                    codigo_keys = ['codigo', 'codigo/nombre', 'nombre', 'codigo_identificador']
                    codigo = next((row_data.get(k,'').strip() for k in codigo_keys if row_data.get(k,'')), '')
                    if not codigo:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: código vacío para NAP."); continue
                    if codigo in existing_naps:
                        omitidos+=1; continue
                    puertos_keys = ['cantidad_puertos', 'cantidad_puertos(nap)', 'puertos']
                    puertos_val = str(next((row_data.get(k,'') for k in puertos_keys if row_data.get(k,'')), '8'))
                    sector_obj = Sectores.objects.first()  # asignar sector por defecto
                    CajasNap.objects.create(codigo_identificador=codigo, latitud=lat, longitud=lng, cantidad_puertos=puertos_val, sector=sector_obj)
                    existing_naps.add(codigo); importados+=1

                elif tipo == 'HUB':
                    codigo_keys = ['codigo', 'codigo/nombre', 'nombre', 'codigo_identificador']
                    codigo = next((row_data.get(k,'').strip() for k in codigo_keys if row_data.get(k,'')), '')
                    if not codigo:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: código vacío para HUB."); continue
                    if codigo in existing_hubs:
                        omitidos+=1; continue
                    Hubs.objects.create(codigo=codigo, latitud=lat, longitud=lng)
                    existing_hubs.add(codigo); importados+=1

                elif tipo == 'SEDE':
                    nombre_keys = ['nombre', 'codigo/nombre', 'codigo', 'codigo_identificador']
                    nombre = next((row_data.get(k,'').strip() for k in nombre_keys if row_data.get(k,'')), '')
                    if not nombre:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: nombre vacío para SEDE."); continue
                    if nombre in existing_sedes:
                        omitidos+=1; continue
                    desc_keys = ['descripcion', 'descripcion(sede)']
                    descripcion = next((row_data.get(k,'') for k in desc_keys if row_data.get(k,'')), '')
                    Sedes.objects.create(nombre=nombre, descripcion=descripcion, latitud=lat, longitud=lng)
                    existing_sedes.add(nombre); importados+=1

                elif tipo == 'SECTOR':
                    codigo_keys = ['codigo', 'codigo/nombre', 'nombre', 'codigo_identificador']
                    codigo = next((row_data.get(k,'').strip() for k in codigo_keys if row_data.get(k,'')), '')
                    prefijo_keys = ['prefijo', 'prefijo(sector)']
                    prefijo = next((row_data.get(k,'').strip() for k in prefijo_keys if row_data.get(k,'')), '')
                    if not codigo or not prefijo:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: código o prefijo vacío para SECTOR."); continue
                    if codigo in existing_sectores:
                        omitidos+=1; continue
                    lat_c_keys = ['latitud_centro', 'lat_centro', 'latitud', 'lat', 'lat_inicio', 'latitud/lat_inicio']
                    lng_c_keys = ['longitud_centro', 'lng_centro', 'longitud', 'lng', 'lng_inicio', 'longitud/lng_inicio']
                    lat_c = float(next((row_data.get(k,'') for k in lat_c_keys if row_data.get(k,'')), lat) or lat)
                    lng_c = float(next((row_data.get(k,'') for k in lng_c_keys if row_data.get(k,'')), lng) or lng)
                    Sectores.objects.create(nombre=codigo, prefijo_comercial=prefijo, latitud_centro=lat_c, longitud_centro=lng_c)
                    existing_sectores.add(codigo); importados+=1

                elif tipo == 'FIBRA':
                    codigo_keys = ['codigo', 'codigo/nombre', 'nombre', 'codigo_identificador']
                    codigo = next((row_data.get(k,'').strip() for k in codigo_keys if row_data.get(k,'')), '')
                    if not codigo:
                        errores+=1; detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: código vacío para FIBRA."); continue
                    if codigo in existing_fibras:
                        omitidos+=1; continue
                    lat_f_keys = ['lat_fin', 'latitud_fin', 'lat_fin(solo fibra)', 'latitud_fin(solo fibra)']
                    lng_f_keys = ['lng_fin', 'longitud_fin', 'lng_fin(solo fibra)', 'longitud_fin(solo fibra)']
                    lat_f = float(next((row_data.get(k,'') for k in lat_f_keys if row_data.get(k,'')), lat) or lat)
                    lng_f = float(next((row_data.get(k,'') for k in lng_f_keys if row_data.get(k,'')), lng) or lng)
                    buffers_keys = ['cantidad_buffers', 'cantidad_buffers(fibra)', 'buffers']
                    hilos_keys = ['hilos_por_buffer', 'hilos_por_buffer(fibra)', 'hilos']
                    buffers = int(next((row_data.get(k,'') for k in buffers_keys if row_data.get(k,'')), 4) or 4)
                    hilos = int(next((row_data.get(k,'') for k in hilos_keys if row_data.get(k,'')), 12) or 12)
                    FibrasOpticas.objects.create(
                        codigo_identificador=codigo, latitud_inicio=lat, longitud_inicio=lng,
                        latitud_fin=lat_f, longitud_fin=lng_f,
                        cantidad_buffers=buffers, hilos_por_buffer=hilos, capacidad_total=buffers*hilos
                    )
                    existing_fibras.add(codigo); importados+=1

            except Exception as e:
                errores+=1
                detalle_errores.append(f"Hoja '{sheet_name}' fila {row_idx}: {str(e)[:80]}")

    return {
        'importados':    importados,
        'omitidos':      omitidos,
        'errores':       errores,
        'detalle_errores': detalle_errores[:20],  # máximo 20 mensajes de error
    }
